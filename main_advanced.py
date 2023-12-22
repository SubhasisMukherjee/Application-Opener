import subprocess
import pygetwindow as gw
import os
import webbrowser
import time
import win32com.client
import openpyxl
from config.config import DEBUG


def find_latest_file_or_dir(path):
    max_id = 0
    latest_item = ""
    latest_item_name = ""

    for pth in os.listdir(path):
        id = int(pth.split('.')[0])
        if id > max_id:
            max_id = id
            latest_item = pth.strip()
            latest_item_name = pth.split('.')[1].strip()

    return max_id, latest_item, latest_item_name



def create_course_map(tracker_path):
    workbook = openpyxl.load_workbook(tracker_path)
    sheet = workbook['Sheet1']
    dict_course_map = {}

    for row in range(2, sheet.max_row+1):
        if sheet.cell(row=row, column=1).fill.start_color.index == 5:
            course_section = sheet.cell(row=row, column=1).value.strip()
            dict_course_map[course_section] = []
        elif sheet.cell(row=row, column=1).value is not None:
            dict_course_map[course_section].append((sheet.cell(row=row, column=1).value.strip(),
                                                    sheet.cell(row=row, column=3).value.strip() if sheet.cell(row=row, column=3).value is not None else sheet.cell(row=row, column=3).value))

    return dict_course_map

# Task 1: Execute mongosh.exe
try:
    subprocess.Popen(['start', 'cmd', '/k', "mongosh.exe"], shell=True)
    time.sleep(2)   # to cater to the time delay needed to open cmd window
    cmd_window = gw.getActiveWindow()
    cmd_window.maximize()
    # # Create STARTUPINFO structure to configure the window size
    # startupinfo = subprocess.STARTUPINFO()
    # startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    # startupinfo.wShowWindow = ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 3)  # SW_MAXIMIZE

except FileNotFoundError:
    print("mongosh.exe not found. Please check the path or install MongoDB.")


# Task 2: Open an excel spreadsheet in MS Excel
course_tracker_path = r"C:\Users\Subhasis\Desktop\Udemy - MongoDB .xlsx"
try:
    # Open the file with excel
    subprocess.run(["start", "excel", course_tracker_path], shell=True)
    excel_window = gw.getActiveWindow()
    excel_window.maximize()

except Exception as e:
    print(f"Error: {e}")


# Task 3: Open the latest directory in File Explorer
try:
    dir_path = "D:\\Study\\Data Engineering\\Udemy\\MongoDB - The Complete Developer's Guide 2023\\"
    # max_subdir_idx, max_file_idx = 0, 0
    # latest_subdir, latest_section, latest_file, latest_lecture, l_lectures_current_section = "", "", "", "", ""
    max_subdir_idx, latest_subdir, latest_section = find_latest_file_or_dir(dir_path)
    latest_subdir_path = dir_path + latest_subdir
    max_file_idx, latest_file, latest_lecture = find_latest_file_or_dir(latest_subdir_path)
    if DEBUG == "Y":
        print(f"Latest subdirectory: {latest_subdir}")
        print(f"Latest section: {latest_section}")
        print(f"Latest file: {latest_file}")
        print(f"Latest lecture: {latest_lecture}")

    d_course_map = create_course_map(course_tracker_path)
    if DEBUG == "Y":
        print(d_course_map)

    l_sections = list(d_course_map.keys())
    l_lectures_current_section = d_course_map[latest_section]
    if DEBUG == "Y":
        print(f"Total list of sections: {l_sections}")
        print(f"Lectures in current section: {l_lectures_current_section}")

    new_lecture_file_to_open = False
    new_section_dir_to_open = False

    for idx in range(len(l_lectures_current_section)):
        lecture = l_lectures_current_section[idx][0]
        status = l_lectures_current_section[idx][1]
        last_lecture_current_section = l_lectures_current_section[-1][0]
        if DEBUG == "Y":
            print(f"Lecture: {lecture}, Status: {status}")

        if lecture == latest_lecture and status == "Complete" and lecture != last_lecture_current_section:
            new_lecture_file_to_open = True
            new_lecture_to_start = l_lectures_current_section[idx+1][0]
            section_dir_to_open_abs_path = dir_path + latest_subdir
            lecture_file_to_open_abs_path = section_dir_to_open_abs_path + "\\" + str(max_file_idx+1) + ". " + new_lecture_to_start + ".docx"
            if DEBUG == "Y":
                print("Old section continuing, new lecture to start")
                print(f"New lecture to start: {new_lecture_to_start}")
                print(f"Subdirectory to open: {section_dir_to_open_abs_path}")
                print(f"File to open: {lecture_file_to_open_abs_path}")
            break
        elif status is None:
            section_dir_to_open_abs_path = dir_path + latest_subdir
            lecture_file_to_open_abs_path = section_dir_to_open_abs_path + "\\" + latest_file
            if DEBUG == "Y":
                print("Old section and lecture continuing")
                print(f"Subdirectory to open: {section_dir_to_open_abs_path}")
                print(f"File to open: {lecture_file_to_open_abs_path}")
            break

        elif lecture == last_lecture_current_section:
            new_section_dir_to_open = True
            new_lecture_file_to_open = True
            last_section_idx = l_sections.index(latest_section)
            new_section_to_start = l_sections[last_section_idx+1]
            new_lecture_to_start = d_course_map[new_section_to_start][0][0]
            section_dir_to_open_abs_path = dir_path + str(max_subdir_idx+1) + ". " + new_section_to_start
            lecture_file_to_open_abs_path = section_dir_to_open_abs_path + "\\1. " + new_lecture_to_start + ".docx"
            if DEBUG == "Y":
                print("New section and lecture to start")
                print(f"New section to start: {new_section_to_start}")
                print(f"New lecture to start: {new_lecture_to_start}")
                print(f"Subdirectory to open: {section_dir_to_open_abs_path}")
                print(f"File to open: {lecture_file_to_open_abs_path}")

            break

    if new_section_dir_to_open == True:
        os.mkdir(section_dir_to_open_abs_path)

    subprocess.Popen("explorer " + section_dir_to_open_abs_path)
    time.sleep(2)
    cmd_window = gw.getActiveWindow()
    cmd_window.maximize()

    # Create a new instance of the Word application
    word_app = win32com.client.Dispatch("Word.Application")
    # Make Word visible (optional)
    word_app.Visible = True
    # Create a new document if needed
    if new_lecture_file_to_open == True:
        doc = word_app.Documents.Add()
        doc.SaveAs(lecture_file_to_open_abs_path)
    # Open the Word document
    doc = word_app.Documents.Open(lecture_file_to_open_abs_path)

except FileNotFoundError:
    print(f"Directory not found: {latest_subdir_path}")

except ValueError:
    print(f"The naming format of the directory of the file is not correct")

except Exception as excp:
    print(f"Error occured: {excp}")


# Task 4: Open a specific URL in Chrome
chrome_url = "https://www.udemy.com/course/mongodb-the-complete-developers-guide/"
webbrowser.open(chrome_url)

# # Check if the specific part of the URL is already open in Chrome
# url_is_open = False
# # for _ in range(3):  # Try a few times (adjust as needed)
# #    webbrowser.open(chrome_url)
# #    time.sleep(2)  # Wait for Chrome to open
# open_chrome_processes = [p.name() for p in psutil.process_iter(['name']) if 'chrome' in p.info['name'].lower()]
# if any(chrome_url in process for process in open_chrome_processes):
#     url_is_open = True


# Task 5: Open a text file in Notepad++
notepad_plus_plus_path = r"C:\Program Files\Notepad++\notepad++.exe"
file_to_open = r"C:/Users/Subhasis/Desktop/MongoDB  commands.txt"

try:
    subprocess.Popen([notepad_plus_plus_path, file_to_open])
except FileNotFoundError:
    print("Notepad++ not found. Please check the path to Notepad++ executable.")
except Exception as e:
    print(f"An error occurred: {e}")


## Task 6:  Open MS Paint
try:
    subprocess.Popen([r"C:\Users\Subhasis\AppData\Local\Microsoft\WindowsApps\mspaint.exe"])
except FileNotFoundError:
    print("MS Paint not found. Please check the path to MS Paint executable.")
except Exception as e:
    print(f"An error occurred: {e}")


## Task 7: open Snipping Tool
try:
    subprocess.Popen([r"C:\Users\Subhasis\AppData\Local\Microsoft\WindowsApps\SnippingTool.exe"])
except FileNotFoundError:
    print("Snipping Tool not found. Please check the path to Snipping Tool executable.")
except Exception as e:
    print(f"An error occurred: {e}")
